#!/usr/bin/env python3

import boto3
from boto3.dynamodb.conditions import Key
from botocore.exceptions import ClientError
import json
import yaml
import os
import time
# from datetime import datetime
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

# from antiope_scorecard.common import *
# try:
#     from lib.account import *
#     from lib.common import *
# except ImportError as e:
#     print("Cannot find the Antiope Libraries")
#     print("Error: {}".format(e))
#     exit(1)


import logging
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
# Quiet Boto3
logging.getLogger('botocore').setLevel(logging.WARNING)
logging.getLogger('boto3').setLevel(logging.WARNING)
logging.getLogger('urllib3').setLevel(logging.WARNING)


headers = ["accountId", "requirementId", "resourceId", "reason", "expirationDate", "hidden", "approval", "Type"]

def main(args):

    wb = load_workbook(filename=args.filename, read_only=True)
    ws = wb["exceptions"]

    dynamodb = boto3.resource('dynamodb')
    exclusions_table = dynamodb.Table(args.table_name)

    ids = []
    for cell in ws[1]:
        ids.append(cell.value)

    for row in ws.iter_rows(row_offset=1):

        row_item={}

        if type(row[0].value) is int:
            account_id = str(row[0].value).zfill(12)
        else:
            account_id = str(row[0].value)

        for cell in row:
            if cell.value is None:
                continue
            key = ids[cell.column -1]
            logger.debug(f"key {key} for {account_id}/{row[1].value} is {cell.value}")
            row_item[key] = cell.value

        # Sanity check the XLS
        skip_row = False
        for h in headers:
            if h not in row_item:
                logger.error(f"row {row} is missing key {h}")
                skip_row = True
                break;

        # Ignore this row if all the attributes aren't there
        if skip_row:
            continue

        exclusion_id = f"{row_item['requirementId']}#{row_item['resourceId']}"
        ddb_item = {
            'requirementId':        row_item['requirementId'],
            'accountId':            f"{account_id}",
            'rqrmntId_rsrceRegex':  exclusion_id,
            'formFields':           {},
            "resourceId":           row_item['resourceId'],
            'type':                 row_item['Type'],
            'status':               "approved",
            'hidesResources':       row_item['hidden'],
            'lastStatusChangeDate': str(datetime.datetime.now()),
            'adminComments':        "Copied from old-scorecard exclusions_table",
            'lastModifiedByAdmin':  row_item['approval']
        }
        ddb_item['formFields']['reason'] = row_item['reason']

        if isinstance(row_item['expirationDate'], datetime.datetime):
            ddb_item['expirationDate'] = row_item['expirationDate'].strftime("%Y/%m/%d")
        else:
            ddb_item['expirationDate'] = row_item['expirationDate']

        logger.debug(f"Adding Exclusion for {exclusion_id} in {account_id} Data:\n {json.dumps(ddb_item, sort_keys=True, indent=2)}")
        try:
            response = exclusions_table.put_item(Item=ddb_item)
        except ClientError as e:
            logger.critical(f"Unable to save exclusion {account_id}/{exclusion_id}: {e}")
            raise

def do_args():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug", help="print debugging info", action='store_true')
    parser.add_argument("--error", help="print error info only", action='store_true')
    parser.add_argument("--filename", help="Excel File to import", required=True)
    parser.add_argument("--table-name", help="DDB Table to Import into", required=True)

    args = parser.parse_args()

    # Logging idea stolen from: https://docs.python.org/3/howto/logging.html#configuring-logging
    # create console handler and set level to debug
    ch = logging.StreamHandler()
    if args.debug:
        logger.setLevel(logging.DEBUG)
    elif args.error:
        logger.setLevel(logging.ERROR)
    else:
        logger.setLevel(logging.INFO)
    # create formatter
    # formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    formatter = logging.Formatter('%(name)s - %(levelname)s - %(message)s')
    # add formatter to ch
    ch.setFormatter(formatter)
    # add ch to logger
    logger.addHandler(ch)

    # if not hasattr(args, 'environment_id'):
    #     print("Must specify --environment_id")
    #     exit(1)

    return(args)


if __name__ == '__main__':
    args = do_args()
    main(args)