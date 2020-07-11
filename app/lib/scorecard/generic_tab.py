from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .generic_formatter import Formatter

def create_tab(worksheet: Worksheet, rows: list, formatting: Formatter):
    """
    Function to specifically create the Non Compliant Resource worksheet. Modified passed in workbook.

    Parameters:
    worksheet (Worksheet): The worksheet to modify.
    rows (list): Data for the rows of each item
    formatting (object): Object containing various formatting information
    """

    # add header
    worksheet.append(formatting.get_header_names())

    # sort data, since adding a sort to the filter has no effect until excel sorts it
    sort_header = ([h for h in formatting.get_headers() if h.sort] + [None])[0]
    if sort_header:
        rows = list(sorted(rows, key=sort_header.get_value))

    for row in rows:
        worksheet.append(formatting.format_resource(row))

    # no footer

    worksheet.title = formatting.title
    worksheet.freeze_panes = formatting.freeze

    # add filtering capability
    if formatting.excel_filter:
        worksheet.auto_filter.ref = 'A1:{}{:d}'.format(
            get_column_letter(len(formatting.get_header_names())),
            len(rows) + 1 # one header row + number of rows
        )

    # set column widths
    for idx, header in enumerate(formatting.get_headers()):
        worksheet.column_dimensions[get_column_letter(idx + 1)].width = header.width

    # set column conditional formatting
    for idx, header in enumerate(formatting.get_headers()):
        if header.conditional_formatting:
            column_range = '{0}2:{0}{1:d}'.format(
                get_column_letter(idx + 1), # excel is 0 based indexing
                len(rows) + 1 # number of rows + 1 header row
            )
            worksheet.conditional_formatting.add(column_range, header.conditional_formatting)

    # bold header row
    for header_cell in worksheet[1]:
        header_cell.font = Font(bold=True)

    return worksheet
