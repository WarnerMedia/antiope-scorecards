"""Functions and formatting for generating the matrix tab (scores by requirement and account)"""
from datetime import datetime

from openpyxl.formatting.rule import  Rule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from lib.dynamodb import config_table, scores_table

class MatrixTabFormatting():
    TITLE = 'Scores by Account, Itemized'
    FREEZE = 'D3'
    # Column headers before account columns
    HEADERS = [
        'Description',
        'Requirement ID',
        'Severity'
    ]
    SEVERITY_COLUMN = 3 # excel columns are 1 based
    ACCOUNT_SCORE = 'Account score'
    DEFAULT_SEVERITY = 'ok'
    def __init__(self):
        # These configs must be fetched at runtime
        self.severity_colors = config_table.get_config(config_table.SEVERITYCOLORS)
        self.severity_weights = config_table.get_config(config_table.SEVERITYWEIGHTS)
        self.version = config_table.get_config(config_table.VERSION)
        self.severity_formatting = self._create_severity_formatting()
        self.error_formatting = self._create_error_formatting()

    def excel_string(self, value):
        return f'"{value}"'

    def _create_error_formatting(self):
        return [
            {
                'value': 'Err', # TODO make this a constant
                'description': 'Error generating spreadsheet',
                'fill': '111111',
                'font_color': 'FF0000',
                'font_bold': True,
                'font_italics': True
            },
            {
                'value': scores_table.DATA_NOT_COLLECTED,
                'description': 'Data Not Collected',
                'fill': '0000D0', # Blue
                'font_color': 'FFFFFF',
                'font_bold': True,
                'font_italics': True
            },
            {
                'value': scores_table.NOT_APPLICABLE,
                'description': 'Score does not apply for account.',
                'fill': self.severity_formatting[self.DEFAULT_SEVERITY]['fill'],
                'font_color': self.severity_formatting[self.DEFAULT_SEVERITY]['font_color'],
                'font_bold': False,
                'font_italics': False
            },
            {
                'value': 0, # TODO make this a constant
                'description': 'No failing resources',
                'fill': self.severity_formatting[self.DEFAULT_SEVERITY]['fill'],
                'font_color': self.severity_formatting[self.DEFAULT_SEVERITY]['font_color'],
                'font_bold': False,
                'font_italics': False
            },
        ]

    def _create_severity_formatting(self):
        severity_formatting = {}
        for severity, colors in self.severity_colors.items():
            severity_formatting[severity] = {
                'weight': self.severity_weights.get(severity, -1),
                'fill': colors['background'],
                'font_color': colors['text'],
            }
        severity_formatting[self.DEFAULT_SEVERITY] = {
            'weight': -1,
            'fill': self.severity_colors[self.DEFAULT_SEVERITY]['background'],
            'font_color': self.severity_colors[self.DEFAULT_SEVERITY]['text'],
        }
        return severity_formatting

def create_matrix_tab(
        worksheet: Worksheet,
        matrix_rows: list,
        account_overall_scores: dict,
        accounts: dict
    ) -> Worksheet:
    """
    Function to generate the workbook based data already gatered and parsed.

    Parameters:
    matrix_rows (list): Direct input for the itemized worksheet.
    account_overall_scores (dict): Mapping from account id to account overall score
    accounts (list): List of accounts.

    Returns:
    Workbook: The workbook object ready to be saved.
    """
    formatting = MatrixTabFormatting()

    ### Add data ###

    # header rows
    account_header = []
    for account in accounts:
        if 'account_name' in account:
            account_header.append(account['account_name'])
        else:
            account_header.append(account['accountId'])
    # add header row
    worksheet.append(formatting.HEADERS + account_header)

    # add account score rows
    worksheet.append([formatting.ACCOUNT_SCORE, '', ''] + list(account_overall_scores.values()))

    # add requirement rows
    rows = sorted(matrix_rows, key=lambda row: row['description']) # sort by description field
    for row in rows:
        worksheet.append([
            row['description'], row['requirementId'], row['severity']
        ] + row['numFailing'])
        if all(score == scores_table.NOT_APPLICABLE for score in row['numFailing']):
            worksheet.row_dimensions[worksheet.max_row].hidden = True

    # add footer
    worksheet.append(['']) # empty row
    worksheet.append([f'Scored Against CSS Version: {formatting.version}'])
    worksheet.append([f'Report Generated at {datetime.now()} GMT'])

    ### Apply formatting ###

    worksheet.title = formatting.TITLE

    # bold headers
    for header_cell in worksheet[1][:len(formatting.HEADERS)]:
        header_cell.font = Font(bold=True, size=11)

    # vertically align account names for readability
    for account_name in worksheet[1][len(formatting.HEADERS):]:
        account_name.alignment = Alignment(text_rotation=45)

    # word wrap long descriptions
    for description in worksheet['A']:
        description.alignment = Alignment(wrap_text=True)

    # freeze first column and first row
    worksheet.freeze_panes = formatting.FREEZE

    # bold overall scores
    overall_score_row = 2
    for grade_cell in worksheet[overall_score_row][:worksheet.max_column]:
        grade_cell.font = Font(bold=True, size=11)

    # right align ACCOUNT_SCORE cell
    worksheet[overall_score_row][1].alignment = Alignment(horizontal='right')

    # set appropriate font size
    for row in worksheet.iter_rows(min_row=overall_score_row + 1):
        for cell in row:
            cell.font = Font(size=9)

    # set Description column width
    worksheet.column_dimensions['A'].width = 80

    # set other column widths
    for col_index in range(len(formatting.HEADERS) + 1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = 8


    # hide requirement id column
    worksheet.column_dimensions['B'].hidden = True

    # cell coloring/conditional formatting
    # format account scores
    colors_ordered_by_weight = list(reversed(sorted(formatting.severity_formatting.values(), key=lambda severity: severity['weight'])))

    for account_score in worksheet[overall_score_row][len(formatting.HEADERS):]:
        try:
            score = int(account_score.value)
        except: # pylint: disable=bare-except
            score = 0
        account_score.number_format = '0'
        # colors in reverse order by weight so first one encountered is correct
        for color in colors_ordered_by_weight:
            if score >= color['weight']:
                account_score.fill = PatternFill(start_color=color['fill'], end_color=color['fill'], fill_type='solid')
                account_score.font = Font(color=color['font_color'], bold=True)

                break

    # add conditional formatting for error scores
    score_cell_range = '{}3:{}{:d}'.format(
        get_column_letter(len(formatting.HEADERS) + 1),
        get_column_letter(worksheet.max_column),
        len(matrix_rows) + 2
    )
    score_cell_top_left = '{}3'.format(get_column_letter(len(formatting.HEADERS) + 1))

    for error_format in formatting.error_formatting:
        # convert python string to excel string
        if isinstance(error_format['value'], str):
            check_value = formatting.excel_string(error_format['value'])
        else:
            check_value = error_format['value']
        worksheet.conditional_formatting.add(
            score_cell_range,
            Rule(
                type='expression',
                formula=[f'{check_value}={score_cell_top_left}'],
                priority=worksheet.conditional_formatting.max_priority + 1,
                stopIfTrue=True,
                dxf=DifferentialStyle(
                    font=Font(
                        color=error_format['font_color']
                    ),
                    fill=PatternFill(
                        start_color=error_format['fill'],
                        end_color=error_format['fill'],
                        fill_type='solid',
                    )
                )
            )
        )

    severity_column_reference = '${}3'.format(get_column_letter(formatting.SEVERITY_COLUMN))

    for severity, severity_format in formatting.severity_formatting.items():
        # convert python string to excel string
        check_value = formatting.excel_string(severity)
        worksheet.conditional_formatting.add(
            score_cell_range,
            Rule(
                type='expression',
                formula=[f'{check_value}={severity_column_reference}'],
                priority=worksheet.conditional_formatting.max_priority + 1,
                stopIfTrue=True,
                dxf=DifferentialStyle(
                    font=Font(
                        color=severity_format['font_color']
                    ),
                    fill=PatternFill(
                        start_color=severity_format['fill'],
                        end_color=severity_format['fill'],
                        fill_type='solid',
                    )
                )
            )
        )

    return worksheet
