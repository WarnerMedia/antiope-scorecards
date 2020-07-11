from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from lib.dynamodb import config_table


class NcrTabFormatting():
    TITLE = 'Non Compliant Resources'
    FREEZE = 'A2'

    VALID_EXCLUSION_FONT = Font(italic=True)
    INVALID_EXCLUSION_FONT = Font(italic=True, color='FF0000') # red text
    EXCLUSION_HEADER_PREFIX = 'Exclusion '
    FIXED_HEADERS = [
        {
            'headerText': 'Account ID',
            'width': 14,
            'key': ['accountId'],
        },
        {
            'headerText': 'Account Name',
            'width': 42,
            'key': ['accountName'],
        },
        {
            'headerText': 'Resource ID',
            'width': 22,
            'key': ['resourceId'],
        },
        {
            'headerText': 'Resource Type',
            'width': 15,
            'key': ['combinedServiceComponent'],
        },
        {
            'headerText': 'Region',
            'width': 12,
            'key': ['region'],
        },
        {
            'headerText': 'Severity',
            'width': 9,
            'key': ['severity'],
        },
        {
            'headerText': 'Requirement Violated',
            'width': 80,
            'key': ['description'],
            'sort': True
        },
        {
            'headerText': 'Reason',
            'width': 80,
            'key': ['reason'],
        },
        {
            'headerText': 'Requirement ID',
            'width': 21,
            'key': ['requirementId'],
        },
        {
            'headerText': 'Exclusion Admin Comments',
            'width': 30,
            'key': ['exclusion.adminComments'],
        },
        {
            'headerText': 'Exclusion Expiration Date',
            'width': 21,
            'key': ['exclusionExpiration'],
        },
        {
            'headerText': 'Expiration Applied',
            'width': 16,
            'key': ['exclusionApplied'],
            'exclusionApplied': True
        }
    ]

    def __init__(self):
        self.exclusion_types = config_table.get_config(config_table.EXCLUSIONS)
        self._add_exclusion_headers()

    def get_exclusion_applied_header_index(self):
        for idx, header in enumerate(self.headers):
            if header.get('exclusionApplied'):
                return idx
        raise TypeError('No exclusion applied header found')

    def _add_exclusion_headers(self):
        unique_exclusion_fields = {}
        for exclusion_type in self.exclusion_types.values():
            for field_name, field_definition in exclusion_type.get('formFields', {}).items():
                if field_definition.get('showInNcrView', False):

                    header_name = '{}{}'.format(self.EXCLUSION_HEADER_PREFIX, field_definition['label'])
                    header_key = 'exclusion.formFields.{}'.format(field_name)

                    if header_name in unique_exclusion_fields: # only add a header once (fields can appear in multiple exclusion types)
                        # exclusions can have different fieldNames but the same label. We consolidate by label.
                        if header_key not in unique_exclusion_fields[header_name]['key']:
                            unique_exclusion_fields[header_name]['key'].append(header_key)
                    else:
                        unique_exclusion_fields[header_name] = {
                            'headerText': header_name,
                            'width': 20,
                            'key': [header_key],
                        }
        self.headers = self.FIXED_HEADERS + list(unique_exclusion_fields.values())

    def get_header_names(self):
        return [header['headerText'] for header in self.headers]

    def format_resource(self, ncr: dict) -> list:
        """Extract fields from NCR based on headers"""

        output = []
        for header in self.headers:
            output.append(get_value(header, ncr))

        return output


def create_ncr_tab(worksheet: Worksheet, ncr_data: list):
    """
    Function to specifically create the Non Compliant Resource worksheet. Modified passed in workbook.

    Parameters:
    worksheet (Worksheet): The worksheet to modify.
    ncr_data (list): The ncr data to be dropped into the worksheet.
    """

    formatting = NcrTabFormatting()
    # add header
    worksheet.append(formatting.get_header_names())

    # sort data, since adding a sort to the filter has no effect until excel sorts it
    sort_header = ([h for h in formatting.headers if h.get('sort')] + [None])[0]
    if sort_header:
        ncrs = sorted(ncr_data, key=lambda ncr: get_value(sort_header, ncr))
    else:
        ncrs = ncr_data

    for ncr in ncrs:
        if ncr.get('isHidden'): # We've marked this one for hiding
            continue
        worksheet.append(formatting.format_resource(ncr))

    # no footer

    worksheet.title = formatting.TITLE
    worksheet.freeze_panes = formatting.FREEZE

    # starting_column = ord('A') + scorecard.NCR_STARTING_COLUMN
    for idx, header in enumerate(formatting.headers):
        worksheet.column_dimensions[get_column_letter(idx + 1)].width = header['width']

    # bold header row
    for header_cell in worksheet[1]:
        header_cell.font = Font(bold=True)
    starting_column = 'A'
    # add filtering capability
    worksheet.auto_filter.ref = 'A1:{}{:d}'.format(
        get_column_letter(worksheet.max_column),
        worksheet.max_row
    )

    italics_max_row = max(worksheet.max_row, 3)

    # add conditional formatting for resource rows with a valid exclusion (italics)
    cell_range = '{}2:{}{:d}'.format(starting_column, get_column_letter(worksheet.max_column), italics_max_row)
    exclusion_valid_column = get_column_letter(formatting.get_exclusion_applied_header_index() + 1)
    worksheet.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=['AND(${0}2=TRUE, NOT(ISBLANK(${0}2)))'.format(exclusion_valid_column)],
            font=formatting.VALID_EXCLUSION_FONT
        )
    )

    # add conditional formatting for resource rows with an expired exclusion (italic red text)
    worksheet.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=['AND(${0}2=FALSE, NOT(ISBLANK(${0}2)))'.format(exclusion_valid_column)],
            font=formatting.INVALID_EXCLUSION_FONT,
        )
    )

    return worksheet


def get_value(header: dict, resource: dict) -> str:
    """
    Abstraction method to allow header names for the NCR worksheet to be mapped and generalized in a static value.
    This allows you to easily adjust the look and positioning of the NCR worksheet without changing much of any code.

    Parameters:
    header (dict): The header definition to use to get the data.
    resource (dict): The resource which is being parsed for mapping.

    Returns:
    str: The found value to be put into the list.
    """

    for key in header['key']:
        keys = key.split('.')
        value = get_value_by_path(resource, keys)
        if value != '':
            return value
    return ''

def get_value_by_path(dictionary: dict, keys: list):
    for key in keys:
        if isinstance(dictionary, dict):
            dictionary = dictionary.get(key, '')
        else:
            return ''
    return dictionary
