from collections import defaultdict
from copy import deepcopy
import decimal
import io
import json
import os
from enum import Enum, auto
from datetime import date
from typing import Dict, List, NewType, Optional, Tuple

from boto3.dynamodb.conditions import Key
from openpyxl import Workbook

from lib.dynamodb import accounts_table, ncr_table, requirements_table, scans_table, scores_table, user_table
from lib.dynamodb.scores import ScoresTable
from lib.lambda_decorator.decorator import states_decorator
from lib.logger import logger
from lib.s3.s3_buckets import S3
from lib.scorecard import matrix_tab, ncr_tab, accounts_tab, sponsors_tab

RequirementId = NewType('RequirementId', str)
AccountId = NewType('AccountId', str)
Scores = Dict[AccountId, Dict[RequirementId, Dict]]

DATA_NOT_COLLECTED = ScoresTable.DATA_NOT_COLLECTED
BUCKET = os.environ.get('SCORECARD_BUCKET')
PREFIX = os.environ.get('SCORECARD_PREFIX')

SPONSOR_FIELD = 'exec_sponsor_email'
NO_SPONSOR_VALUE = 'No executive sponsor'
class SheetTypes(Enum):
    SINGLE_ACCOUNT = auto()
    PAYER_ACCOUNT = auto()
    USER = auto()
    GLOBAL = auto()


CACHE: Dict[str, Dict] = {}
CACHED_SCAN = {
    'scanId': None
}

@states_decorator
def gen_spreadsheets_handler(event, context):
    scan_id = event['openScan']['scanId']

    # clear CACHE if the scan has changed
    if CACHED_SCAN['scanId'] != scan_id:
        logger.debug('Clearing cache')
        CACHE.clear()
        CACHED_SCAN['scanId'] = scan_id

    accounts, s3_key, sheet_type = get_accounts(event)

    if not accounts:
        logger.info('No accounts, nothing to do')
        return

    load_scores(accounts)
    scores = get_scores()
    requirements = get_requirements()
    ncr_data = get_ncr(scan_id, accounts, sheet_type)

    workbook = create_base_workbook(ncr_data, accounts, requirements, scores)

    if sheet_type in [SheetTypes.GLOBAL, SheetTypes.PAYER_ACCOUNT]:
        add_accounts_tab(workbook, accounts, scores)
        add_sponsor_tab(workbook, accounts, scores)

    if sheet_type == SheetTypes.GLOBAL:
        # write date stamped global spreadsheet
        write_to_s3(workbook, '{}/global/scorecard-{}.xlsx'.format(PREFIX, date.today()))

        logger.debug('Writing global json scores')
        score_export = create_score_export(scan_id, accounts, scores, requirements)
        write_global_json_scores(score_export)

        logger.debug('Writing resource json')
        resource_export = create_resource_export(ncr_data, requirements)
        write_global_resources(resource_export)

    logger.debug('Writing to s3')
    write_to_s3(workbook, s3_key)

def add_accounts_tab(workbook: Workbook, accounts: List, scores: Scores):
    accounts_worksheet = workbook.create_sheet()
    account_overall_scores = build_overall_score(accounts, scores)
    rows = []
    for account in accounts:
        rows.append({
            'accountId': account['accountId'],
            'account_name': account['account_name'],
            'score': account_overall_scores[account['accountId']]
        })

    accounts_tab.create_accounts_tab(accounts_worksheet, rows)

def add_sponsor_tab(workbook: Workbook, accounts: List, scores: Scores):
    accounts_worksheet = workbook.create_sheet()
    account_overall_scores = build_overall_score(accounts, scores)
    rows = defaultdict(lambda: {'executiveSponsor':'', 'accountCount': 0, 'sumOfScores': 0})

    for account in accounts:
        if SPONSOR_FIELD in account and isinstance(account[SPONSOR_FIELD], str):
            sponsor = account[SPONSOR_FIELD].lower()
        else:
            sponsor = NO_SPONSOR_VALUE
        rows[sponsor]['executiveSponsor'] = sponsor
        rows[sponsor]['accountCount'] += 1
        rows[sponsor]['sumOfScores'] += account_overall_scores[account['accountId']]

    sponsors_tab.create_sponsors_tab(accounts_worksheet, rows.values())

def create_base_workbook(ncr_data: List, accounts: List, requirements: dict, scores: Scores) -> Workbook:
    account_overall_scores = build_overall_score(accounts, scores)
    matrix_rows = prepare_requirement_scores(accounts, scores, requirements)

    # join requirements to ncrs
    ncr_rows = []

    for original_ncr in ncr_data:
        new_ncr = deepcopy(original_ncr)
        new_ncr.update(requirements[new_ncr['requirementId']])
        new_ncr['combinedServiceComponent'] = f'{new_ncr.get("service", "")} {new_ncr.get("component", "")}'
        ncr_rows.append(new_ncr)
    logger.debug('Preparing workbooks')
    # create excel spreadsheet
    workbook = Workbook()

    # add matrix (account summary tab)
    matrix_worksheet = workbook.active
    matrix_tab.create_matrix_tab(matrix_worksheet, matrix_rows, account_overall_scores, accounts)

    # add NCR sheet
    ncr_worksheet = workbook.create_sheet()
    ncr_tab.create_ncr_tab(ncr_worksheet, ncr_rows)

    return workbook

#############################
#####Gather Generic Data#####
#############################
def get_accounts(payload: dict) -> Tuple[List, str, SheetTypes]:
    """
    Parses inbound event for account specification. Options are:
    accountId, payerId, userEmail, or global. If nothing is found, global is default.

     Parameters:
     payload (dict): the full event to be parsed for account specification.

     Returns:
     List: a list of to generate the spreadsheet for
     str: a string representing the s3 key to write the spreadsheet to
     str: type of spreadsheet being created
    """
    logger.debug('Looking up accounts to process')
    if 'accountId' in payload:
        prefix = '{}/by-account/{}.xlsx'.format(PREFIX, payload['accountId'])
        accounts = [get_account(payload['accountId'])]
        return accounts, prefix, SheetTypes.SINGLE_ACCOUNT
    elif 'payerId' in payload:
        prefix = '{}/by-payer/{}.xlsx'.format(PREFIX, payload['payerId'])
        accounts = [a for a in get_all_accounts().values() if a.get('payer_id') == payload['payerId']]
        return accounts, prefix, SheetTypes.PAYER_ACCOUNT
    elif 'userEmail' in payload:
        prefix = '{}/by-user/{}.xlsx'.format(PREFIX, payload['userEmail'])
        user = user_table.get_user(payload['userEmail'])
        accounts = []
        for account_id in user.get('accounts', {}).keys():
            account = get_account(account_id)
            if account:
                accounts.append(account)
        return accounts, prefix, SheetTypes.USER
    else:
        return list(get_all_accounts().values()), '{}/global/scorecard-latest.xlsx'.format(PREFIX), SheetTypes.GLOBAL


###---DDB QUERY---###
def get_all_accounts() -> Dict[str, Dict]:
    """
    Wrapper for ddb query to getting all accounts. This also loads into cache for future lookups.

    Returns:
    dict: accounts indexed by account id.
    """
    if 'accounts' not in CACHE:
        logger.debug('Getting accounts for cache')
        CACHE['accounts'] = { # index accounts by account id
            account['accountId']: account for account in accounts_table.scan_all()
        }
    return CACHE['accounts']

def get_account(account_id) -> Optional[dict]:
    """Get account by account id from cache"""
    accounts = get_all_accounts()
    logger.debug('Getting account ID: %s', account_id)
    return accounts.get(account_id, {})

###---DDB QUERY---###
def get_requirements() -> Dict[RequirementId, Dict]:
    """
    Saves requirements into a cache for future use. Returns requirements indexed by requirement_id
    """
    if 'requirements' not in CACHE:
        logger.debug('Getting requirements for cache')
        CACHE['requirements'] = {
            req['requirementId']: req for req in requirements_table.scan_all() if not req.get('ignore', False)
        }
    return CACHE['requirements']

###---DDB QUERY---###
def get_scores() -> Scores:
    """Get account scores dictonary from cache"""
    return CACHE['account_detail_scores']

def load_scores(accounts: List) -> None:
    """
    Queries accounts' scores from database and stores in CACHE

    Parameters:
    accounts (list): the account to get the scores for.
    """
    scan_id = CACHED_SCAN['scanId']
    if 'account_detail_scores' not in CACHE:
        CACHE['account_detail_scores'] = defaultdict(lambda: defaultdict(dict))
    for account in accounts:
        account_id = account['accountId']
        if account_id not in CACHE['account_detail_scores']:
            logger.debug('Querying account scores')
            account_scores = scores_table.query_all(
                KeyConditionExpression=Key('scanId').eq(scan_id) & Key('accntId_rqrmntId').begins_with(account_id)
            )
            CACHE['account_detail_scores'][account_id] = defaultdict(dict, {
                score['requirementId']: score for score in account_scores
            })

def create_score_export(scan_id: str, accounts: List, scores: Scores, requirements) -> Dict:
    account_overall_scores = build_overall_score(accounts, scores)
    account_json_list = []
    for account in accounts:
        account_scores = scores[account['account_id']]
        account_json_list.append({
            'accountId': account['accountId'],
            'accountName': account.get('account_name') or account.get('accountId'),
            'overallScore': account_overall_scores[account['accountId']],
            'requirementScores': {
                requirement_id: account_scores[requirement_id] for requirement_id in requirements.keys()
            }
        })

    json_scores = {
        'scores': account_json_list,
        'scanId': scan_id
    }

    return json_scores

def create_resource_export(ncrs: list, requirements: dict):
    export = []

    for original_ncr in ncrs:
        ncr = deepcopy(original_ncr)
        # remove dynamodb composite keys
        ncr.pop('accntId_rsrceId_rqrmntId', None)
        ncr.pop('rqrmntId_accntId', None)
        ncr['requirement'] = requirements[ncr['requirementId']]
        export.append(ncr)
    return export

##############################
#####Prepare Summary Data#####
##############################
def prepare_requirement_scores(accounts: list, scores: Scores, requirements: dict) -> list:
    """
    Prepares data for direct use by the scorecard for import into the itemized worksheet.

    Parameters:
    accounts (list): list of accounts to report scores for

    Returns:
    list: a list of dictionary which corresponds directly for reporting requirements. Structure of dicts is:
        {
            description: requirement description
            severity: requirement severity
            source: requirement source
            numFailing: [score, score, score] # one score per account in order
        }
    """
    logger.debug('Preparing initial data')
    matrix_rows = []
    for requirement_id, requirement in requirements.items():
        row = {
            'description': requirement['description'],
            'requirementId': requirement['requirementId'],
            'severity': requirement['severity'],
            'numFailing': [],
        }

        for account in accounts:
            try:
                detailed_score = scores[account['accountId']][requirement_id]
                num_failing = next(iter(detailed_score['score'].values()))['numFailing']
                row['numFailing'].append(num_failing)
            except: # pylint: disable=bare-except
                # add 'Err' if the score doesn't exist for some reason (it should always be created)
                row['numFailing'].append('Err')
        matrix_rows.append(row)
    return matrix_rows

def build_overall_score(accounts: List, scores: Scores) -> Dict[str, int]:
    """
    Creates overall account score for a list of accounts

    Returns dictionary mapping from account number to score
    """
    account_overall_scores = {}
    for account in accounts:
        account_scores = scores[account['accountId']].values()
        aggregated_score = scores_table.weighted_score_aggregate_calc(account_scores)
        account_overall_scores[account['accountId']] = scores_table.get_single_score_calc(aggregated_score)
    return account_overall_scores

###---DDB QUERY---###
def get_ncr(scan_id: str, accounts: list, sheet_type: SheetTypes) -> list:
    """
    Wrapper function for ddb queries to get ncr data.

    Parameters:
    scan_id (str): The id for the scan to lookup NCR data for.
    accounts (list): All accounts for this spreadsheet.
    global_scan (bool): Boolean value representing whether this is a global scan or not.

    Returns:
    list: List of NCR records.
    """

    logger.debug('Getting NCRs')
    scan_results = []
    if sheet_type == SheetTypes.GLOBAL:
        scan_results = ncr_table.query_all(
            KeyConditionExpression=Key('scanId').eq(scan_id)
        )
    else:
        for account in accounts:
            results = ncr_table.query_all(
                KeyConditionExpression=Key('scanId').eq(scan_id) &
                Key('accntId_rsrceId_rqrmntId').begins_with(account['accountId'])
            )
            scan_results.extend(results)
    return scan_results

######################
#####Persist Data#####
######################
def write_to_s3(workbook: Workbook, s3_key: str):
    """
    Function to save workbook to s3.

    Parameters:
    workbook (Workbook): The workbook to be saved.
    s3_key (str): If not a global scan, this will be used for the prefix. Otherwise ignored if global scan.
    """
    logger.debug('Writing spreadsheet to s3://%s/%s', BUCKET, s3_key)
    write_stream = io.BytesIO()
    workbook.save(write_stream)
    workbook_bytes = write_stream.getbuffer().tobytes()
    if os.getenv('WRITE_LOCAL'):
        logger.debug('Writing to local disk, not uploading to s3')
        with open('local.xlsx', 'wb') as local_file_xlsx:
            local_file_xlsx.write(workbook_bytes)
        return

    S3.put_object(
        Bucket=BUCKET,
        Key=s3_key,
        Body=workbook_bytes
    )

def write_global_resources(ncrs: list):
    json_string = json.dumps(decimal_to_num(ncrs), indent=2)
    if os.getenv('WRITE_LOCAL'):
        logger.debug('Writing to local disk, not uploading to s3')
        with open('resources.local.json', 'w') as local_file:
            local_file.write(json_string)
        return

    S3.put_object(
        Bucket=BUCKET,
        Key='{}/global/resources-{}.json'.format(PREFIX, date.today()),
        Body=json_string
    )
    S3.put_object(
        Bucket=BUCKET,
        Key='{}/global/resources-latest.json'.format(PREFIX),
        Body=json_string
    )

def write_global_json_scores(json_scores: dict):
    json_string = json.dumps(decimal_to_num(json_scores), indent=2)

    if os.getenv('WRITE_LOCAL'):
        logger.debug('Writing to local disk, not uploading to s3')
        with open('scores.local.json', 'w') as local_file:
            local_file.write(json_string)
        return

    S3.put_object(
        Bucket=BUCKET,
        Key='{}/global/scorecard-{}.json'.format(PREFIX, date.today()),
        Body=json_string
    )
    S3.put_object(
        Bucket=BUCKET,
        Key='{}/global/scorecard-latest.json'.format(PREFIX),
        Body=json_string
    )


def decimal_to_num(obj):
    """
    Helper function to convert all decimal valued inputs to the real representation of the value (int or float.)
    This function is recursive.

    Parameters:
    obj (obj): An object to parse for decimals.

    Returns:
    obj: The passed in object with any transformations made.
    """
    if isinstance(obj, list):
        for item in range(len(obj)): # pylint: disable=consider-using-enumerate
            obj[item] = decimal_to_num(obj[item])
    elif isinstance(obj, dict):
        for key, value in obj.items():
            obj[key] = decimal_to_num(value)
    elif isinstance(obj, decimal.Decimal):
        if obj % 1 == 0:
            obj = int(obj)
        else:
            obj = float(obj)
    return obj

@states_decorator
def gen_spreadsheets_error_handler(event, context):
    scans_table.add_error(event['openScan']['scanId'], context.function_name, event['error'])

@states_decorator
def setup_user_spreadsheets_handler(event, context):
    """
    Adds user emails to event/state so step functions can
    iterator over them to generate spreadsheets

    Expected input event format
    {
        "scanId": scan_id,
    }

    Returns {
        "scanId": scan_id,
        "userEmails": list of user's email addresses
    }
    """
    # add usersEmails to event/state
    users = user_table.scan_all()
    event['userEmails'] = [user['email'] for user in users]
    # remove account numbers to keep payload size down
    event.pop('load', None)
    return event
